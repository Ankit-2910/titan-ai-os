"""
TITAN File Generator Tool
Generates downloadable Excel/CSV files based on user requests.
Returns base64-encoded file content for frontend download.
"""
import io
import csv
import json
import base64
from typing import Dict, Any, List
from app.tools.registry import BaseTool


class FileGeneratorTool(BaseTool):

    @property
    def name(self) -> str:
        return "generate_file"

    @property
    def description(self) -> str:
        return (
            "Generate downloadable files (Excel/CSV) for the user. "
            "Use this when user asks for templates, rosters, schedules, "
            "trackers, reports, or any structured data they want to download."
        )

    @property
    def parameters(self) -> Dict[str, Any]:
        return {
            "file_type": {
                "type": "string",
                "description": "Type of file: 'csv' or 'excel'",
            },
            "filename": {
                "type": "string",
                "description": "Name for the file (without extension)",
            },
            "headers": {
                "type": "array",
                "description": "List of column headers",
            },
            "rows": {
                "type": "array",
                "description": "List of rows, each row is a list of values",
            },
            "description": {
                "type": "string",
                "description": "Brief description of what this file contains",
            },
        }

    async def execute(
        self,
        file_type: str,
        filename: str,
        headers: List[str],
        rows: List[List[Any]],
        description: str = "",
    ) -> Dict[str, Any]:
        try:
            if file_type.lower() == "csv":
                return self._generate_csv(filename, headers, rows, description)
            elif file_type.lower() in ("excel", "xlsx"):
                return self._generate_excel(filename, headers, rows, description)
            else:
                return {"success": False, "result": None, "error": f"Unsupported: {file_type}"}
        except Exception as e:
            return {"success": False, "result": None, "error": str(e)}

    def _generate_csv(self, filename, headers, rows, description):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        csv_content = output.getvalue()
        encoded = base64.b64encode(csv_content.encode("utf-8")).decode("utf-8")
        return {
            "success": True,
            "result": {
                "file_type": "csv",
                "filename": f"{filename}.csv",
                "content_base64": encoded,
                "description": description,
                "rows_count": len(rows),
                "mime_type": "text/csv",
            },
            "error": None,
        }

    def _generate_excel(self, filename, headers, rows, description):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = filename[:31]

            # Header style — TITAN purple
            header_fill = PatternFill("solid", fgColor="4C1D95")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin = Side(style="thin", color="E0E0E0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Write rows
            alt_fill = PatternFill("solid", fgColor="F5F0FF")
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

            # Auto-width columns
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save to buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            encoded = base64.b64encode(buf.read()).decode("utf-8")

            return {
                "success": True,
                "result": {
                    "file_type": "excel",
                    "filename": f"{filename}.xlsx",
                    "content_base64": encoded,
                    "description": description,
                    "rows_count": len(rows),
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                "error": None,
            }
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return self._generate_csv(filename, headers, rows, description)
